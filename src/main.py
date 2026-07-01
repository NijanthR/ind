from __future__ import annotations

if __package__ is None or __package__ == "":
    import sys
    from pathlib import Path

    sys.path.append(str(Path(__file__).resolve().parents[1]))

import argparse
import csv
import logging
from pathlib import Path
from typing import Optional

from .config import DEFAULT_OUTPUT_DIR, DEFAULT_OUTPUT_NAME, DEFAULT_REPORT_NAME, DEFAULT_TOP_N

DEFAULT_EXCEL_NAME = "ranked_candidates.xlsx"
from .embedder import SemanticEmbedder
from .ingestion import (
    build_candidate_text,
    discover_dataset_root,
    load_candidate_schema,
    load_jsonl_candidates,
    load_job_description,
    load_sample_candidates,
    load_sample_submission,
    parse_job_description_features,
)
from .report import generate_report
from .scorer import cheap_prefilter_score, normalize_and_sort, score_candidate


LOGGER = logging.getLogger(__name__)


def _print_dataset_inspection(dataset_root: Path) -> None:
    schema = load_candidate_schema(dataset_root)
    sample_candidates = load_sample_candidates(dataset_root)
    sample_submission = load_sample_submission(dataset_root)
    job_text = load_job_description(dataset_root)

    LOGGER.info("Dataset root: %s", dataset_root)
    LOGGER.info("Candidate schema required fields: %s", ", ".join(schema.get("required", [])))
    if sample_candidates:
        first = sample_candidates[0]
        LOGGER.info("Sample candidate columns: %s", ", ".join(sorted(first.keys())))
        for row in sample_candidates[:3]:
            profile = row.get("profile", {}) or {}
            LOGGER.info(
                "Sample row: %s | %s | %s years | %s",
                row.get("candidate_id"),
                profile.get("anonymized_name"),
                profile.get("years_of_experience"),
                profile.get("current_title"),
            )
    if sample_submission:
        LOGGER.info("Sample submission columns: %s", ", ".join(sample_submission[0].keys()))
        LOGGER.info("Sample submission first row: %s", sample_submission[0])
    LOGGER.info("Job description characters: %s", len(job_text))


def _load_candidate_records(dataset_root: Path, candidates_path: Path | None) -> Path:
    if candidates_path is not None:
        return candidates_path
    found = list(dataset_root.rglob("candidates.jsonl"))
    if not found:
        raise FileNotFoundError("candidates.jsonl was not found under the dataset root")
    return found[0]


def _write_csv(output_path: Path, candidates: list) -> None:
    with output_path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.writer(handle)
        writer.writerow(["candidate_id", "rank", "score", "reasoning"])
        for rank, candidate in enumerate(candidates[:DEFAULT_TOP_N], start=1):
            writer.writerow([
                candidate.candidate_id,
                rank,
                f"{candidate.final_score / 100.0:.6f}",
                candidate.reasoning,
            ])


def _write_excel(output_path: Path, candidates: list) -> None:
    """Write a richly formatted Excel workbook with all score components."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        LOGGER.warning("openpyxl not installed — skipping Excel output. Run: pip install openpyxl")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Ranked Candidates"

    # ── Header styling ──────────────────────────────────────────────────────
    header_fill  = PatternFill("solid", fgColor="1F3864")
    header_font  = Font(bold=True, color="FFFFFF", size=11)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=False)
    thin_side    = Side(style="thin", color="CCCCCC")
    thin_border  = Border(left=thin_side, right=thin_side, bottom=thin_side)

    # Tier colour map
    tier_fills = {
        "Strong Match":    PatternFill("solid", fgColor="C6EFCE"),  # green
        "Good Match":      PatternFill("solid", fgColor="FFEB9C"),  # yellow
        "Moderate Match":  PatternFill("solid", fgColor="FFCC99"),  # orange
        "Weak Match":      PatternFill("solid", fgColor="FFC7CE"),  # red
    }

    headers = [
        "Rank", "Candidate ID", "Name", "Current Title", "Years Exp.",
        "Final Score", "Tier",
        "Semantic", "Skills", "Experience", "Education",
        "Platform Activity", "Career Trajectory", "Anomaly Penalty",
        "Reasoning",
    ]

    col_widths = [
        6, 16, 22, 26, 11,
        13, 18,
        10, 10, 12, 12,
        18, 18, 16,
        60,
    ]

    for col_idx, (header, width) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font      = header_font
        cell.fill      = header_fill
        cell.alignment = center_align
        cell.border    = thin_border
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"  # freeze header row

    # ── Data rows ───────────────────────────────────────────────────────────
    for rank, c in enumerate(candidates[:DEFAULT_TOP_N], start=1):
        row = [
            rank,
            c.candidate_id,
            c.candidate_name,
            c.current_title,
            round(c.years_of_experience, 1),
            round(c.final_score / 100.0, 6),
            c.recommendation_tier,
            round(c.semantic_score, 2),
            round(c.skills_match_score, 2),
            round(c.experience_score, 2),
            round(c.education_score, 2),
            round(c.platform_activity_score, 2),
            round(c.career_trajectory_score, 2),
            round(c.anomaly_penalty, 2),
            c.reasoning,
        ]
        for col_idx, value in enumerate(row, start=1):
            cell = ws.cell(row=rank + 1, column=col_idx, value=value)
            cell.alignment = Alignment(vertical="center", wrap_text=(col_idx == len(row)))
            cell.border    = thin_border

        # colour-code the Tier cell (col 7)
        tier_cell = ws.cell(row=rank + 1, column=7)
        fill = tier_fills.get(c.recommendation_tier)
        if fill:
            tier_cell.fill = fill

        # alternate row shading
        if rank % 2 == 0:
            row_fill = PatternFill("solid", fgColor="F5F7FA")
            for col_idx in range(1, len(headers) + 1):
                cell = ws.cell(row=rank + 1, column=col_idx)
                if not cell.fill or cell.fill.fgColor.rgb in ("00000000", "FFFFFFFF"):
                    cell.fill = row_fill

    # ── Auto-filter on header row ───────────────────────────────────────────
    ws.auto_filter.ref = ws.dimensions

    wb.save(output_path)
    LOGGER.info("Excel report saved → %s", output_path)


def run_pipeline(dataset_root: Path, candidates_path: Path | None, output_dir: Path, excel: bool = False) -> list:
    job_text = load_job_description(dataset_root)
    jd_features = parse_job_description_features(job_text)
    candidate_file = _load_candidate_records(dataset_root, candidates_path)
    embedder = SemanticEmbedder()

    LOGGER.info("Scoring candidates from %s", candidate_file)
    pref_candidates = []
    for candidate in load_jsonl_candidates(candidate_file):
        prefilter = cheap_prefilter_score(candidate, jd_features)
        candidate["_prefilter_score"] = prefilter
        pref_candidates.append(candidate)

    pref_candidates.sort(key=lambda item: (-float(item.get("_prefilter_score", 0.0)), str(item.get("candidate_id", ""))))
    shortlist_size = min(len(pref_candidates), 100)
    shortlisted = pref_candidates[:shortlist_size]

    scored_candidates = []
    batch_items = []
    batch_texts = []
    batch_size = 128 if embedder.available else 256
    for candidate in shortlisted:
        profile_text = build_candidate_text(candidate)
        candidate["_profile_text"] = profile_text
        batch_items.append(candidate)
        batch_texts.append(profile_text)
        if len(batch_items) >= batch_size:
            similarities = embedder.batch_similarities(jd_features.raw_text, batch_texts)
            for item, similarity in zip(batch_items, similarities):
                scored_candidates.append(score_candidate(item, jd_features, similarity))
            batch_items = []
            batch_texts = []
    if batch_items:
        similarities = embedder.batch_similarities(jd_features.raw_text, batch_texts)
        for item, similarity in zip(batch_items, similarities):
            scored_candidates.append(score_candidate(item, jd_features, similarity))

    scored_candidates = normalize_and_sort(scored_candidates)

    output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = output_dir / DEFAULT_OUTPUT_NAME
    report_path = output_dir / DEFAULT_REPORT_NAME

    _write_csv(csv_path, scored_candidates)
    generate_report(report_path, scored_candidates)

    if excel:
        excel_path = output_dir / DEFAULT_EXCEL_NAME
        _write_excel(excel_path, scored_candidates)

    return scored_candidates


def build_parser() -> argparse.ArgumentParser:
    parser = argparse.ArgumentParser(description="AI candidate ranking pipeline")
    parser.add_argument("--dataset-root", type=Path, default=None, help="Path to the extracted challenge directory")
    parser.add_argument("--candidates", type=Path, default=None, help="Optional path to candidates.jsonl")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Output directory")
    parser.add_argument("--excel", action="store_true", default=False, help="Also write output/ranked_candidates.xlsx")
    return parser


def main() -> None:
    logging.basicConfig(level=logging.INFO, format="%(levelname)s %(name)s: %(message)s")
    args = build_parser().parse_args()
    dataset_root = args.dataset_root or discover_dataset_root()
    _print_dataset_inspection(dataset_root)
    scored_candidates = run_pipeline(dataset_root, args.candidates, args.output_dir, excel=args.excel)
    LOGGER.info("Wrote %s and %s", args.output_dir / DEFAULT_OUTPUT_NAME, args.output_dir / DEFAULT_REPORT_NAME)
    if args.excel:
        LOGGER.info("Excel  → %s", args.output_dir / DEFAULT_EXCEL_NAME)
    for rank, candidate in enumerate(scored_candidates[:10], start=1):
        LOGGER.info(
            "Top %s: %s | %s | %.2f | %s",
            rank,
            candidate.candidate_id,
            candidate.candidate_name,
            candidate.final_score,
            candidate.recommendation_tier,
        )


if __name__ == "__main__":
    main()
